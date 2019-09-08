import datetime
import itertools
import pandas as pd
import re

from copy import copy
from gsheets import Sheets
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

class GSheets(object):
    """ Class to manage Google Sheets documents

    Attributes
    ----------
    wb: gsheets.models.SpreadSheet
        Workbook including all projects
    sh_name: string, default 'BDD Missions'
        Name of the sheet name with all projects
    ws: gsheets.models.WorkSheet
        Worksheet with all projects
    """
    def __init__(self, url, sh_name='Form Responses 1',
                 credentials='from_drive_to_xlsx/config/credentials.json',
                 storage_json='from_drive_to_xlsx/config/storage.json'):
        """
        Parameters
        ----------
        url: string
            URL to access the workbook
        sh_name: string, default 'BDD Missions'
            Name of the sheet name with all projects
        credentials: string, default 'config/credentials_gsheets.json'
            Path to access the JSON credentials file
        storage_json: string, default 'config/storage.json'
            Path to access the JSON storage file
        """
        workbook = Sheets.from_files(credentials, storage_json).get(url)
        self.wb = workbook

        self.sh_name = sh_name
        self.ws = workbook.find(sh_name)

    def get_data(self):
        """ Get and clean data from worksheet

        Returns
        -------
        df: pd.DataFrame
        """
        df = self.ws.to_frame()

        df.columns = [re.sub(' |\?', '', re.sub('é', 'e', col)).lower() for col in df.columns]

        # df['timestamp'] = pd.to_datetime(df['timestamp'])
        df['date'] = pd.to_datetime(df['date'], format='%m/%d/%Y')
        # df['date'] = df['date'].apply(lambda x: '{}/{}/{}'.format(x.day, x.month, x.year))

        return df

    def get_data_from_timestamp(self, timestamp):
        """ Collect data with filter from a specific timestamp

        Parameters
        ----------
        timestamp: datetime.datetime

        Returns
        -------
        df: pd.DataFrame
        """
        df = self.get_data()

        df['timestamp'] = pd.to_datetime(df['timestamp'])
        df.query('timestamp > "{}"'.format(timestamp), inplace=True)

        df['timestamp'] = df['timestamp'].apply(str)

        return df

    def get_number_of_rows_to_append(self, spendings, account_name):
        """ Return the number of rows to add to Excel file

        Parameters
        ----------
        spendings: pd.DataFrame
            Information collected through Google Form
        account_name: string
            Account name

        Returns
        -------
        rows_to_append: int
            Number of rows to add
        """
        account_spendings = spendings.query('compte == "{}"'.format(account_name))

        rows_to_append = account_spendings.shape[0]

        return rows_to_append

    def get_timestamp_and_values(self, spendings, account_name):
        """ Get timestamps values and spendings information

        Parameters
        ----------
        spendings: pd.DataFrame
            Information collected through Google Form
        account_name: string
            Account name

        Returns
        -------
        account_timestamp: list
            Timestamp values
        values: list
            Spendings information to add
        """
        account_spendings = spendings.query('compte == "{}"'.format(account_name))

        account_timestamp = account_spendings['timestamp'].values

        list_row_values = list(dataframe_to_rows(account_spendings.drop(labels=['timestamp', 'compte'], axis=1),
                                                 index=False, header=False))
        values = itertools.chain(*list_row_values)

        return account_timestamp, values


class ExcelWorkbook(object):
    """ Class to manage Excel workbook

    Attributes
    ----------
    workbook: openpyxl.Workbook
    """
    def __init__(self, filename):
        """
        Parameters
        ----------
        filename: string
            Path to access Excel file
        """
        self.workbook = load_workbook(filename=filename)

    def get_latest_timestamp(self):
        """ Select latest timestamp in all account sheets
        """
        latest_timestamp = datetime.datetime(2017, 1, 1)

        def _convert_value_to_datetime(cell):
            try:
                return datetime.datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S')
            except:
                return 'Not a date'

        for ws in self.workbook.sheetnames:
            if ws not in ['Récap', 'Catégories']:
                col_timestamp = self.workbook[ws]['A']
                values_col = list(map(_convert_value_to_datetime, col_timestamp))

                try:
                    max_value = max([x for x in values_col if isinstance(x, datetime.datetime)])
                    latest_timestamp = max(latest_timestamp, max_value)
                except:
                    # print('No timestamp in {}'.format(ws))
                    pass

        return latest_timestamp

    def get_last_row_number(self, ws_name, start_row=12, column='C'):
        """ Select last row with information to append Excel sheet

        Parameters
        ----------
        ws_name: string
            Sheetname to investigate
        start_row: integer, default 12
            Row to start from for investigation
        column: string, default 'C'
            Column to check values - usually, column C contains mandatory field

        Returns
        -------
        row: integer
            First row with empty cell value
        """
        ws = self.workbook[ws_name]

        value = 'to_start'
        row = start_row
        while value:
            row += 1
            value = ws['{}{}'.format(column, row)].value

        return row

    def insert_and_copy_format(self, ws_name, row_to_insert_from, number_rows):
        """ Insert number_rows in Excel sheet starting from row_to_insert_from
        and copy format of last row to all inserted rows

        Parameters
        ----------
        ws_name: string
            Sheetname to investigate
        row_to_insert_from: integer
            Insert rows from row #row_to_insert_from - most often, it corresponds
            to ExcelWorkbook.get_last_row_number(...)
        number_rows: integer
            Number of rows to insert
        """
        ws = self.workbook[ws_name]
        latest_row = ws[row_to_insert_from - 1]

        ws.insert_rows(idx=row_to_insert_from, amount=number_rows)

        for i in range(row_to_insert_from, row_to_insert_from+number_rows+1):
            inserted_row = ws[i]
            for inserted_cell, last_row_cell in zip(inserted_row, latest_row):
                inserted_cell.number_format = copy(last_row_cell.number_format)
                inserted_cell.border = copy(last_row_cell.border)

        pass

    def update_total_formula(self, ws_name, row_to_insert_from):
        """ Update total formula from rows that were inserted
        (Insertion using openpyxl does not keep formulas like Excel)

        Parameters
        ----------
        ws_name: string
            Sheetname to investigate
        row_to_insert_from: integer
            Insert rows from row #row_to_insert_from - most often, it corresponds
            to ExcelWorkbook.get_last_row_number(...)
        """
        ws = self.workbook[ws_name]
        last_row = ws.max_row

        if ws_name in ['BNP - CC', 'Amex']:
            cols = ['J', 'I', 'H']
        elif ws_name == 'BNP - Prêt':
            cols = ['H', 'G', 'F']
        else:
            cols = ['I', 'H', 'G']

        for row_number in range(row_to_insert_from, last_row):
            formula_value = '={0}{row_prev} + IF({1}{row}="oui", {2}{row}, 0)'.format(cols[0],
                                                                                      cols[1],
                                                                                      cols[2],
                                                                                      row_prev=row_number-1,
                                                                                      row=row_number)
            ws['{col}{row}'.format(col=cols[0], row=row_number)].value = formula_value

        pass


    def copy_timestamp(self, ws_name, row_start, row_end, list_ts_values):
        """ Copy values of list_ts_values to column A in Excel sheet

        Parameters
        ----------
        ws_name: string
            Sheetname to investigate
        row_start: integer
            Row number to start from
        row_end: integer
            Last row number
        list_ts_values: list
            List of values to copy in column A
        """
        ws = self.workbook[ws_name]

        range_timestamp = ws['A{}:A{}'.format(row_start, row_end)]
        list_range_ts = itertools.chain(*range_timestamp)

        for cell, ts in zip(list_range_ts, list_ts_values):
            cell.value = ts
            cell.font = Font(color='FFFFFF')

        pass

    def copy_spendings_values(self, ws_name, row_start, row_end, list_values):
        """ Copy values of list_values to specific range in Excel sheet

        Parameters
        ----------
        ws_name: string
            Sheetname to investigate
        row_start: integer
            Row number to start from
        row_end: integer
            Last row number
        list_values: list
            Spendings information converted to list
        """
        ws = self.workbook[ws_name]

        if ws_name in ['BNP - CC', 'Amex']:
            range_values = ws['C{}:I{}'.format(row_start, row_end)]
        else:
            range_values = ws['B{}:H{}'.format(row_start, row_end)]

        list_cells = itertools.chain(*range_values)

        for cell, value in zip(list_cells, list_values):
            cell.value = value

        if ws_name in ['BNP - CC', 'Amex']:
            for i in range(row_start, row_end+1):
                ws['B{}'.format(i)].value = ws['C{}'.format(i)].value


if __name__ == "__main__":
    ew = ExcelWorkbook(filename='/Users/vinhpham-gia/Documents/0_Perso/spendings/tmp_test.xlsx')

    # ew.insert_and_copy_format(ws_name='BNP - CC',
    #                           row_to_insert_from=94,
    #                           number_rows=5)

    toto = ew.workbook['BNP - CC']
    cols = ['J', 'I', 'H']
    for row_number in range(94, 667):
        formula_value = '={0}{row_prev} + IF({1}{row}="oui", {2}{row}, 0)'.format(cols[0],
                                                                                  cols[1],
                                                                                  cols[2],
                                                                                  row_prev=row_number-1,
                                                                                  row=row_number)
        toto['{col}{row}'.format(col=cols[0], row=row_number)].value = formula_value


    ew.workbook.save('/Users/vinhpham-gia/Documents/0_Perso/spendings/tmp_test.xlsx')

    pass
